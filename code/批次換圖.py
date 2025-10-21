import streamlit as st
import os
import re
import shutil
import pandas as pd 
import zipfile
import tempfile
from openpyxl import load_workbook
import pythoncom
import win32com.client
from io import BytesIO
from streamlit_extras.stylable_container import stylable_container

def tab5():
    global log
    log = lambda msg: st.warning(msg)

    def modify_source_path(source_path: str, target_path: str) -> tuple[str, str]:
        """
        回傳兩種可能的「已編圖」路徑：
          1. …\\新增圖檔位置已編圖\\JF4558_OK
          2. …\\新增圖檔位置_已編圖\\JF4558_OK
        """
        parent_dir, base_name = os.path.split(source_path)
        dir1 = parent_dir + "已編圖"
        dir2 = parent_dir + "_已編圖"
        cand1 = os.path.join(dir1, f"{base_name}_OK")
        cand2 = os.path.join(dir2, f"{base_name}_OK")
        return cand1, cand2

    def execute_copy_operations(path_mappings: list):
        for mapping in path_mappings:
            src, dst = mapping["source"], mapping["target"]
            if os.path.exists(dst):
                shutil.rmtree(dst)
            if os.path.exists(src):
                shutil.copytree(src, dst)
            else:
                log(f"⚠️ 找不到來源路徑，略過：{src}")

    def rename_and_move_images(base_path, mapping_dict, destination_folder):
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder, exist_ok=True)

        counts = {}

        for i, (article_num, product_id) in enumerate(mapping_dict.items(), start=1):
            folder = os.path.join(base_path, f"{article_num}_OK", "1-Main", "All")
            if not os.path.isdir(folder):
                log(f"⚠️ 找不到資料夾：{folder}")
                counts[article_num] = 0
            else:
                files = [f for f in os.listdir(folder)
                         if os.path.isfile(os.path.join(folder, f))]
                counts[article_num] = len(files)
                for fname in files:
                    m = re.match(rf"^{article_num}_(\d+)\.(.+)$", fname, re.IGNORECASE)
                    if m:
                        seq = int(m.group(1)) - 1
                        ext = m.group(2)
                        new_name = f"{product_id}-{seq}.{ext}"
                        shutil.copy(
                            os.path.join(folder, fname),
                            os.path.join(destination_folder, new_name)
                        )

        return counts

    def convert_xls_to_xlsx(file_path):
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        new_path = file_path.replace('.xls', '.xlsx')
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        wb.SaveAs(os.path.abspath(new_path), FileFormat=51)
        wb.Close()
        excel.Quit()
        pythoncom.CoUninitialize()
        return new_path

    def unprotect_excel_sheet(file_path):
        if file_path.lower().endswith('.xls'):
            file_path = convert_xls_to_xlsx(file_path)
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        for sheet in wb.Sheets:
            if sheet.ProtectContents:
                sheet.Unprotect()
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        pythoncom.CoUninitialize()
        return file_path

    def unzip_and_unprotect(zip_file, extract_to):
        with zipfile.ZipFile(zip_file, 'r') as zf:
            zf.extractall(extract_to)
        unprotected = []
        for fname in os.listdir(extract_to):
            path = os.path.join(extract_to, fname)
            if fname.lower().endswith(('.xls', '.xlsx')):
                try:
                    up = unprotect_excel_sheet(path)
                    unprotected.append(up)
                except Exception as e:
                    log(f"⚠️ 解除保護失敗：{fname}，錯誤：{e}")
        return unprotected

    def handle_submit():
        srcs = [l.strip() for l in st.session_state.source_paths_input.splitlines() if l.strip()]
        tgts = [l.strip() for l in st.session_state.target_paths_input.splitlines() if l.strip()]
        arts = [l.strip() for l in st.session_state.article_nums_input.splitlines() if l.strip()]
        pids = [l.strip() for l in st.session_state.product_ids_input.splitlines() if l.strip()]

        error_flag = (
            not srcs or
            not tgts or
            len(srcs) != len(tgts) or
            not arts or
            not pids or
            len(arts) != len(pids)
        )

        st.session_state['run_srcs'] = srcs
        st.session_state['run_tgts'] = tgts
        st.session_state['run_arts'] = arts
        st.session_state['run_pids'] = pids
        st.session_state['run_error'] = error_flag
        st.session_state['form_submitted'] = True

        if not error_flag:
            st.session_state['source_paths_input'] = ""
            st.session_state['target_paths_input'] = ""
            st.session_state['article_nums_input'] = ""
            st.session_state['product_ids_input'] = ""
            st.session_state['sh_zip_key'] += 1
            st.session_state['y_excel_key'] += 1

    # 初始化 session_state
    for key in (
        "source_paths_input", "target_paths_input",
        "article_nums_input", "product_ids_input",
        "form_submitted", "run_srcs", "run_tgts",
        "run_arts", "run_pids", "run_error"
    ):
        if key not in st.session_state:
            st.session_state[key] = False if key in ("form_submitted","run_error") else ""
    if 'sh_zip_key' not in st.session_state:
        st.session_state['sh_zip_key'] = 0
    if 'y_excel_key' not in st.session_state:
        st.session_state['y_excel_key'] = 0

    mode = st.radio(
        "選擇執行方式 👇",
        ["全圖檔更換", "補情境圖"],
        key="visibility",
        label_visibility="collapsed",
        horizontal=True
    )

    placeholder = st.empty()

    if (not st.session_state['form_submitted']) or st.session_state['run_error']:
        with placeholder.container():
            c1, c2, c3, c4 = st.columns(4)
            c1.text_area("原圖檔位置", key="source_paths_input", height=200)
            c2.text_area("新增圖檔位置", key="target_paths_input", height=200)
            c3.text_area("貨號", key="article_nums_input", height=200)
            c4.text_area("商品頁序號", key="product_ids_input", height=200)

            with st.expander("選填項目", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    with stylable_container(
                        key="col1_file_uploader",
                        css_styles="""
                        {
                          [data-testid='stFileUploaderDropzoneInstructions'] > div > span {
                            display: none;
                          }
                          [data-testid='stFileUploaderDropzoneInstructions'] > div::before {
                            content: '上傳 Shopee 媒體資訊';
                          }
                        }
                        """,
                        ):
                        st.file_uploader(
                            "上傳 Shopee 媒體資訊",
                            type=["zip"],
                            key=f"sh_zip_{st.session_state['sh_zip_key']}",
                            label_visibility="collapsed"
                        )
                with col2:
                    with stylable_container(
                        key="col2_file_uploader",
                        css_styles="""
                        {
                          [data-testid='stFileUploaderDropzoneInstructions'] > div > span {
                            display: none;
                          }
                          [data-testid='stFileUploaderDropzoneInstructions'] > div::before {
                            content: '上傳 Yahoo 圖片資訊';
                          }
                        }
                        """,
                        ):
                        st.file_uploader(
                            "上傳 Yahoo 圖片資訊",
                            type=["xlsx"],
                            key=f"y_excel_{st.session_state['y_excel_key']}",
                            label_visibility="collapsed"
                        )
            st.button(
                "開始執行",
                on_click=handle_submit,
                key="tab4_btn"
            )

            if st.session_state['run_error']:
                srcs = st.session_state['run_srcs']
                tgts = st.session_state['run_tgts']
                arts = st.session_state['run_arts']
                pids = st.session_state['run_pids']
                if not srcs or not tgts:
                    pass
                elif len(srcs) != len(tgts):
                    st.error("來源與目標行數必須相同。")
                elif not arts or not pids:
                    pass
                else:
                    st.error("貨號與商品頁序號行數必須相同。")
                st.session_state['form_submitted'] = False
        return

    placeholder.empty()
    exp_proc = st.expander("執行進度", expanded=True)
    with exp_proc:
        col1, col2 = st.columns(2)
    log = col1.write

    count_dict = {}
    output_files = {}
    collected_images = []
    srcs = st.session_state['run_srcs']
    tgts = st.session_state['run_tgts']
    arts = st.session_state['run_arts']
    pids = st.session_state['run_pids']
    sh_zip = st.session_state.get(f"sh_zip_{st.session_state['sh_zip_key'] - 1}", None)
    y_excel = st.session_state.get(f"y_excel_{st.session_state['y_excel_key'] - 1}", None)

    # ----- 全圖檔更換 邏輯 -----
    if mode == "全圖檔更換":
        col1.write("🔄 正在處理:原圖檔全圖更換")
        processed_srcs = []
        for src in srcs:
            s = src.replace("/", "\\")
            if s.endswith(r"\1-Main\All"):
                s = os.path.dirname(os.path.dirname(s))
            processed_srcs.append(s)

        mappings = []
        all_mod = []
        for s, t in zip(processed_srcs, tgts):
            c1, c2 = modify_source_path(t, s)
            sel = c1 if os.path.exists(c1) else c2
            mappings.append({"source": sel, "target": s})
            all_mod.append(sel)
        execute_copy_operations(mappings)
        col1.write("⭐️ 原圖檔全圖更換完成")

        col1.write("🔄 正在處理:建立官網批次換圖圖片")
        counts = rename_and_move_images(
            os.path.dirname(all_mod[0]),
            dict(zip(arts, pids)),
            r"C:\img_ok"
        )
        if counts:
            col2.dataframe(
                pd.DataFrame(counts.items(), columns=["貨號", "總圖片數"]),
                use_container_width=True,
            )
        count_dict = counts
        col1.write("⭐️ 官網批次換圖圖片格式建立完成")

    # ----- 補情境圖 邏輯 -----
    else:
        col1.write("🔄 正在處理:原圖檔補情境圖")
        count = []
        for src, tgt, art, pid in zip(srcs, tgts, arts, pids):
            c1, c2 = modify_source_path(tgt, src)
            new_base = c1 if os.path.exists(c1) else c2

            new_all = os.path.join(new_base, "1-Main", "All")
            if not os.path.isdir(new_all):
                log(f"⚠️ 找不到新增圖檔資料夾：{new_all}")
                continue
            cnt_new = len([f for f in os.listdir(new_all) if os.path.isfile(os.path.join(new_all, f))])

            orig = src.replace("/", "\\")
            orig_all = (
                orig
                if orig.endswith(r"1-Main\All")
                else os.path.join(orig, "1-Main", "All")
            )
            if not os.path.isdir(orig_all):
                log(f"⚠️ 找不到原圖檔資料夾：{orig_all}")
                continue
            files = [f for f in os.listdir(orig_all) if os.path.isfile(os.path.join(orig_all, f))]

            def seq_of(fn):
                m = re.match(r"^(.*)_(\d+)\.(.+)$", fn)
                return int(m.group(2)) if m else -1

            for fname in sorted(files, key=seq_of, reverse=True):
                m = re.match(r"^(.*)_(\d+)\.(.+)$", fname)
                if not m:
                    continue
                new_seq = int(m.group(2)) + cnt_new
                new_name = f"{m.group(1)}_{new_seq:02d}.{m.group(3)}"
                os.rename(os.path.join(orig_all, fname), os.path.join(orig_all, new_name))

            parent = os.path.dirname(os.path.dirname(orig_all))
            for fname in os.listdir(orig_all):
                m = re.match(rf"^{art}_(\d+)\.(.+)$", fname)
                if m and int(m.group(1)) > 10:
                    shutil.move(os.path.join(orig_all, fname), os.path.join(parent, fname))
            for f in os.listdir(new_all):
                shutil.copy(os.path.join(new_all, f), os.path.join(orig_all, f))
            col1.write("⭐️ 原圖檔補情境圖完成")
            col1.write("🔄 正在處理:建立官網批次換圖圖片")

            dest_ok = r"C:\img_ok"
            os.makedirs(dest_ok, exist_ok=True)
            for fname in os.listdir(orig_all):
                m2 = re.match(rf"^{art}_(\d+)\.(.+)$", fname)
                if m2:
                    new_fn = f"{pid}-{int(m2.group(1)) - 1}.{m2.group(2)}"
                    shutil.copy(os.path.join(orig_all, fname), os.path.join(dest_ok, new_fn))

            total_imgs = len([f for f in os.listdir(orig_all) if os.path.isfile(os.path.join(orig_all, f))])
            count.append((art, total_imgs))

        if count:
            col2.dataframe(pd.DataFrame(count, columns=["貨號", "總圖片數"]), use_container_width=True)
            count_dict = dict(count)
        col1.write("⭐️ 官網批次換圖圖片格式建立完成")

    # ----- 蝦皮批次換圖 Excel -----
    if sh_zip is not None:
        col1.write("🔄 正在處理:蝦皮批次換圖")
        template = r"G:\共用雲端硬碟\TP代營運\1-小工具\自動編圖工具\dependencies\蝦皮批次換圖範本.xlsx"
        tempd = tempfile.mkdtemp()
        try:
            up_files = unzip_and_unprotect(sh_zip, tempd)

            wb = load_workbook(template)
            ws = wb.active
            found = {a: False for a in arts}
            row_i = 7
            for fp in up_files:
                wb2 = load_workbook(fp, data_only=True)
                ws2 = wb2.active
                for row in ws2.iter_rows(values_only=True):
                    artv = row[1]
                    if artv in found and not found[artv]:
                        for ci, val in enumerate(row, start=1):
                            ws.cell(row=row_i, column=ci, value=val)
                        found[artv] = True
                        row_i += 1
                        if all(found.values()):
                            break
                if all(found.values()):
                    break
            for art, ok in found.items():
                if not ok:
                    log(f"⚠️ 未找到貨號：{art}")

            for row in ws.iter_rows(min_row=7, max_row=row_i-1, min_col=5, max_col=16):
                for cell in row:
                    cell.value = None
                    cell.hyperlink = None

            for r in range(7, row_i):
                artv = ws.cell(row=r, column=2).value
                if not artv:
                    continue
                idx = arts.index(artv)
                raw_path = srcs[idx].replace("/", "\\")
                all_folder = (
                    raw_path
                    if raw_path.endswith(r"\1-Main\All")
                    else os.path.join(raw_path, "1-Main", "All")
                )
                if not os.path.isdir(all_folder):
                    log(f"⚠️ 找不到 All 資料夾：{all_folder}")
                    continue

                cnt_total = count_dict.get(artv)
                if cnt_total is None:
                    cnt_total = len([
                        f for f in os.listdir(all_folder)
                        if os.path.isfile(os.path.join(all_folder, f))
                    ])
                cnt = min(cnt_total, 10)
                for i in range(1, cnt + 1):
                    col = 4 + i
                    url = (
                        f"https://tpimage.91app.com/adidas/"
                        f"{artv}_OK/1-Main/All/{artv}_{i:02d}.jpg"
                    )
                    ws.cell(row=r, column=col, value=url)

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            output_files["蝦皮批次換圖.xlsx"] = bio.getvalue()
        finally:
            shutil.rmtree(tempd)
            col1.write("⭐️ 蝦皮批次換圖所需檔案建立完成")

    # ----- Y購批次換圖 Excel & 圖片 -----
    if y_excel is not None:
        col1.write("🔄 正在處理:Y購批次換圖")
        tempd2 = tempfile.mkdtemp()
        try:
            raw_path = os.path.join(tempd2, y_excel.name)
            with open(raw_path, "wb") as f:
                f.write(y_excel.getvalue())

            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            try:
                wb_com = excel.Workbooks.Open(os.path.abspath(raw_path))
                clean_path = raw_path.replace('.xlsx', '_cleaned.xlsx')
                wb_com.SaveAs(os.path.abspath(clean_path), FileFormat=51)
                wb_com.Close(SaveChanges=True)
            except Exception:
                clean_path = raw_path
            finally:
                excel.Quit()
                pythoncom.CoUninitialize()

            try:
                wb_y = load_workbook(clean_path)
            except Exception:
                wb_y = load_workbook(raw_path)

            if "batchImage" in wb_y.sheetnames:
                ws_y = wb_y["batchImage"]
            else:
                for sht in wb_y.worksheets:
                    headers = [cell.value for cell in next(sht.iter_rows(min_row=1, max_row=1))]
                    if "提案人" in headers:
                        ws_y = sht
                        break

            headers = [cell.value for cell in next(ws_y.iter_rows(min_row=1, max_row=1))]
            col_map = {h: i+1 for i, h in enumerate(headers)}

            主圖_cols = [f"屬性項目{i}:商品主圖" for i in range(1,31)]
            圖片_cols = [f"屬性項目{i}:商品圖片" for i in range(1,31)]
            屬性_cols = [f"屬性項目{i}:商品屬性" for i in range(1,31)]

            for art in arts:
                target_row = None
                for r in range(2, ws_y.max_row+1):
                    val = ws_y.cell(row=r, column=col_map.get("賣場名稱",0)).value
                    if val and art in str(val):
                        target_row = r
                        break
                if not target_row:
                    log(f"⚠️ Y購找不到貨號：{art}")
                    continue

                total = count_dict.get(art, 0)

                for title in 主圖_cols + 圖片_cols:
                    idx = col_map.get(title)
                    if idx:
                        ws_y.cell(row=target_row, column=idx).value = None

                主圖_str = f"{art}_01.jpg,{art}_02.jpg"
                for title in 主圖_cols:
                    idx = col_map.get(title)
                    if idx:
                        ws_y.cell(row=target_row, column=idx).value = 主圖_str

                if total > 2:
                    imgs = [f"{art}_{i:02d}.jpg" for i in range(3, total+1)]
                    圖片_str = ",".join(imgs)
                else:
                    圖片_str = ""
                for title in 圖片_cols:
                    idx = col_map.get(title)
                    if idx:
                        ws_y.cell(row=target_row, column=idx).value = 圖片_str

                start_idx = None
                for i, title in enumerate(屬性_cols, start=1):
                    idx = col_map.get(title)
                    if idx and not ws_y.cell(row=target_row, column=idx).value:
                        start_idx = i
                        break
                if start_idx:
                    for i in range(start_idx, 31):
                        t1 = f"屬性項目{i}:商品主圖"
                        t2 = f"屬性項目{i}:商品圖片"
                        for t in (t1, t2):
                            idx = col_map.get(t)
                            if idx:
                                ws_y.cell(row=target_row, column=idx).value = None

                raw_src = srcs[arts.index(art)].replace("/", "\\")
                all_folder = (
                    raw_src
                    if raw_src.endswith(r"\1-Main\All")
                    else os.path.join(raw_src, "1-Main", "All")
                )
                if os.path.isdir(all_folder):
                    for fname in os.listdir(all_folder):
                        if os.path.isfile(os.path.join(all_folder, fname)) and fname.lower() != "thumb.db":
                            collected_images.append((
                                os.path.join(all_folder, fname),
                                f"Y購批次換圖圖片/{fname}"
                            ))
                else:
                    log(f"⚠️ 找不到 All 資料夾：{all_folder}")

            bio_y = BytesIO()
            wb_y.save(bio_y)
            bio_y.seek(0)
            output_files["Y購批次換圖.xlsx"] = bio_y.getvalue()

        finally:
            shutil.rmtree(tempd2)
            col1.write("⭐️ Y購批次換圖所需檔案建立完成")

    # ----- 打包 ZIP 並下載 -----
    if output_files:
        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w") as zf:
            for name, data in output_files.items():
                zf.writestr(name, data)
            for src_path, arcname in collected_images:
                zf.write(src_path, arcname)
        zip_buf.seek(0)
        st.download_button(
            label="下載結果檔",
            data=zip_buf,
            file_name="批次換圖結果.zip",
            mime="application/zip"
        )
    else:
        if st.button("結束任務"):
            st.rerun()

    col1.write("🎉 任務完成")
    st.session_state['form_submitted'] = False
